from openpyxl import Workbook


class ExcelOut:

    def __init__(self, ListValues, outputPath):
        self.path = outputPath
        self.listValues = ListValues




    def writer(self):
        wb = Workbook(guess_types=True)
        ws = wb.active

        row = 1
        col = 1

        lst = [["Project name", self.projectName],
               [None],
               ["Component:", "Variable:", "Value:", "Unit:"]]

        lst.extend([["EOF"]])


        for line in self.listValues:

            for value in line:
                if value is None:
                    pass
                elif type(value) is float:
                    fill = ws.cell(column=col, row=row, value="%g" % value)
                    fill.number_format = '0.00'
                else:
                    _ = ws.cell(column=col, row=row, value="%s" % value)

                col+=1

            row += 1
            col = 1

            wb.save(filename=self.path)

        return "Output file correctly generated"
