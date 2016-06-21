from openpyxl import Workbook
from openpyxl import load_workbook

class Excel:

    def __init__(self, Component, VariableName, Default, filePath):
        self.excelFile = load_workbook(filePath)
        self.sheet = self.excelFile['Sheet1']
        self.component = Component
        self.variableName = VariableName
        self.default = Default


    wb = Workbook()

    def finder(self):
        componentRead = None
        row = 1
        while componentRead != 'EOF':
            valueRead = None
            componentRead = str(self.sheet['A'+str(row)].value)
            if componentRead == self.component:
                while valueRead != self.variableName and valueRead != 'EOC' and row < 1e3:
                    valueRead = str(self.sheet['B'+str(row)].value)
                    row += 1
                if valueRead == self.variableName and self.sheet['C'+str(row-1)].value is not None:
                    print 'LOG:    Value of ' + repr(self.variableName) + ' is set to: ' \
                          + repr(self.sheet['C'+str(row-1)].value)
                    return self.sheet['C'+str(row-1)].value
                elif valueRead == self.variableName and self.sheet['C'+str(row-1)].value is None:
                    print 'LOG:    Could not find value of ' + repr(self.variableName) + '. Using the default value: ' \
                          + repr(self.default)
                    return self.default
                elif valueRead == 'EOC':
                    print 'LOG:    Could not find variable ' + repr(self.variableName) + '. Using the default value: '\
                          + repr(self.default)
                    return self.default
            if componentRead == 'EOF':
                print 'LOG:    Could not find variable ' + repr(self.variableName) + '. Using the default value: '\
                      + repr(self.default)
                return self.default
            row += 1
