from openpyxl import Workbook
from openpyxl import load_workbook

class Excel:

    def __init__(self, Component, VariableName, Default, filePath):
        self.excelFile = load_workbook(filePath)
        self.sheet = self.excelFile['Sheet1']
        self.component = Component
        self.variableName = VariableName
        self.default = Default


    wb = Workbook()

    def finder(self):
        componentRead = None
        row = 1
        while componentRead != 'EOF':
            valueRead = None
            componentRead = str(self.sheet['A'+str(row)].value)
            if componentRead == self.component:
                while valueRead != self.variableName and valueRead != 'EOC' and row < 1e3:
                    valueRead = str(self.sheet['B'+str(row)].value)
                    row += 1
                if valueRead == self.variableName:
                    return self.sheet['C'+str(row-1)].value
                elif valueRead == 'EOC':
                    return self.default
                    print 'Could not find variable ' + str(self.variableName) + '. Using the default value: '
                    + str(self.default)
            if componentRead == 'EOF':
                return self.default
                print 'Could not find variable ' + str(self.variableName) + '. Using the default value: '
                + str(self.default)
            row += 1
