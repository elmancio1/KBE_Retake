from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelImporter:

    filePath= 'C:\Users\Jacopo\Desktop\Academic\GitHub\KBE_Retake\Input\Files\input.xlsx'
    excelFile = load_workbook(filePath)
    sheet = excelFile['Sheet1']

    wb = Workbook()
    componentRead = None
    row = 1

    while componentRead != 'EOF':
        print row
        componentRead = str(sheet['A'+str(row)].value)
        print componentRead
        row += 1




    def fuselageLength(self):
        importValue = float(self.sheet['C5'].value)
        return importValue

    def fuselageRadius(self):
        importValue = float(self.sheet['C6'].value)
        return importValue

    def fuselageRadius2(self):
        importValue = float(self.sheet['C7'].value)
        return importValue
