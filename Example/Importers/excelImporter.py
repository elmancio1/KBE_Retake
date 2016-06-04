from openpyxl import Workbook
from openpyxl import load_workbook

class ExcelImporter:

    def __init__(self, filePath):
        self.filePath= filePath
        self.excelFile = load_workbook(self.filePath)
        self.sheet=self.excelFile['Sheet1']

    wb = Workbook()

    def fuselageLength(self):
        importValue = float(self.sheet['D5'].value)
        return importValue

    def fuselageRadius(self):
        importValue = float(self.sheet['D6'].value)
        return importValue

    def fuselageRadius2(self):
        importValue = float(self.sheet['D7'].value)
        return importValue
